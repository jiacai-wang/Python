from selenium import webdriver
from openpyxl import Workbook
import time


# initiate xlsx
wb = Workbook()
ws = wb.active
ws.title = 'B站' + time.strftime('%Y年%m月%d日',time.localtime()) + 'Top 100'
ws.append(['序号', '标题', '播放量', '弹幕数', 'up主', '网址', '综合得分'])

# get content
url = 'https://www.bilibili.com/ranking'
driver = webdriver.Chrome()
driver.get(url)
listxpath = '//*[@id="app"]/div[1]/div/div[1]/div[2]/div[3]/ul/'

# save ranking into xlsx
for i in range(1, 101):
    itemxpath = listxpath + 'li[' + str(i) + ']'
    ws['A'+ str(i+1)] = driver.find_element_by_xpath(itemxpath + '/div[1]').get_attribute("textContent")
    ws['B'+ str(i+1)] = driver.find_element_by_xpath(itemxpath + '/div[2]/div[2]/a').get_attribute("textContent")
    ws['C'+ str(i+1)] = driver.find_element_by_xpath(itemxpath + '/div[2]/div[2]/div[1]/span[1]').get_attribute("textContent")
    ws['D'+ str(i+1)] = driver.find_element_by_xpath(itemxpath + '/div[2]/div[2]/div[1]/span[2]').get_attribute("textContent")
    ws['E'+ str(i+1)] = driver.find_element_by_xpath(itemxpath + '/div[2]/div[2]/div[1]/a/span').get_attribute("textContent")
    ws['F'+ str(i+1)] = driver.find_element_by_xpath(itemxpath + '/div[2]/div[2]/a').get_attribute("href")
    ws['G'+ str(i+1)] = driver.find_element_by_xpath(itemxpath + '/div[2]/div[2]/div[2]/div').get_attribute("textContent")
wb.save('./top100.xlsx')
driver.close()